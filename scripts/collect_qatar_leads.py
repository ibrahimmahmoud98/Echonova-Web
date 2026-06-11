import os
import re
import sys
import json
import time
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import load_env

EXA_API_KEY = os.environ.get("EXA_API_KEY")
APOLLO_API_KEY = os.environ.get("APOLLO_API_KEY")

EXPORT_DIR = "/Users/suckmyballz/Desktop/ECHONOVA STUDIO/عملاء"
OUTPUT_FILE = os.path.join(EXPORT_DIR, "بينات الشركات المتوسطة القطرية.xlsx")

def log(msg):
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}")

def exa_search(query, num_results=5):
    url = "https://api.exa.ai/search"
    headers = {
        "x-api-key": EXA_API_KEY,
        "content-type": "application/json"
    }
    payload = {
        "query": query,
        "numResults": num_results,
        "useAutoprompt": True
    }
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=15)
        if response.status_code == 200:
            return response.json().get("results", [])
        else:
            log(f"Exa search error {response.status_code}: {response.text}")
    except Exception as e:
        log(f"Exa search exception: {e}")
    return []

def exa_fetch(urls):
    if not urls:
        return []
    url = "https://api.exa.ai/contents"
    headers = {
        "x-api-key": EXA_API_KEY,
        "content-type": "application/json"
    }
    payload = {
        "urls": urls
    }
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=20)
        if response.status_code == 200:
            return response.json().get("results", [])
        else:
            log(f"Exa fetch error {response.status_code}: {response.text}")
    except Exception as e:
        log(f"Exa fetch exception: {e}")
    return []

def scrape_website_direct(domain, website_url):
    log(f"Scraping website directly: {website_url}")
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    emails = []
    whatsapp_numbers = []
    social_links = {}
    
    # Clean up website_url if it doesn't start with http
    if not website_url.startswith("http"):
        website_url = "http://" + website_url
        
    urls_to_try = [website_url]
    if website_url.endswith("/"):
        urls_to_try.append(website_url + "contact")
        urls_to_try.append(website_url + "contact-us")
        urls_to_try.append(website_url + "about")
    else:
        urls_to_try.append(website_url + "/contact")
        urls_to_try.append(website_url + "/contact-us")
        urls_to_try.append(website_url + "/about")
        
    for url in urls_to_try:
        try:
            # Disable SSL verification warnings and verification to handle bad certs
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            if response.status_code == 200:
                html = response.text
                
                # Extract emails
                found_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html)
                for email in found_emails:
                    email_lower = email.lower()
                    # Skip common image extensions or template domains
                    if not any(email_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.css', '.js']):
                        if not any(x in email_lower for x in ['wix', 'wordpress', 'example', 'domain', 'email']):
                            emails.append(email_lower)
                
                # Extract whatsapp links
                wa_matches = re.findall(r'(?:wa\.me|api\.whatsapp\.com/send\?phone=)(\d+)', html)
                for wa in wa_matches:
                    whatsapp_numbers.append(wa)
                    
                # Extract Qatari phone numbers in context of whatsapp (8 digits, starting with 3, 5, 6, 7)
                phone_matches = re.findall(r'(?:\+?974|00974|0)?[3567]\d{7}', html)
                for ph in phone_matches:
                    clean_ph = ph
                    # Standardize Qatari phone numbers to international format: 974XXXXXXXX
                    if ph.startswith(('3', '5', '6', '7')) and len(ph) == 8:
                        clean_ph = '974' + ph
                    elif ph.startswith('0') and len(ph) == 9:
                        clean_ph = '974' + ph[1:]
                    whatsapp_numbers.append(clean_ph)
                    
                # Extract social links
                for platform in ['instagram', 'facebook', 'twitter', 'x.com', 'linkedin', 'youtube', 'tiktok', 'snapchat']:
                    matches = re.findall(fr'https?://(?:www\.)?{platform}\.com/([a-zA-Z0-9_.\-]+)', html)
                    for m in matches:
                        if m not in ['sharer', 'share', 'intent', 'home', 'pages']:
                            social_links[platform] = f"https://{platform}.com/{m}"
        except Exception as e:
            log(f"Failed to fetch {url} directly: {e}")
            
    return list(set(emails)), list(set(whatsapp_numbers)), social_links

def get_marketing_linkedin_profiles(company_name, domain):
    log(f"Searching LinkedIn profiles for marketing managers of: {company_name}")
    query = f"site:linkedin.com/in/ (\"{company_name}\" OR \"{domain}\") (marketing OR advertising OR PR OR communications OR creative OR branding OR \"تسويق\" OR \"دعاية\" OR \"ميديا\")"
    results = exa_search(query, num_results=5)
    profiles = []
    for r in results:
        url = r.get("url", "")
        title = r.get("title", "")
        if "linkedin.com/in/" in url:
            # Try to format the title nicely
            profile_name = title.split("-")[0].strip() if "-" in title else title
            profile_name = profile_name.split("|")[0].strip() if "|" in profile_name else profile_name
            # Exclude generic search result pages
            if not any(x in profile_name.lower() for x in ['linkedin', 'jobs', 'profiles']):
                profiles.append(f"{profile_name}: {url}")
    return profiles[:3]

def enrich_company_details(company):
    name = company['name']
    website_url = company['website_url']
    domain = company['primary_domain']
    
    log(f"Enriching details for: {name} ({domain})")
    
    # 1. Scrape website directly
    emails, whatsapps, socials = scrape_website_direct(domain, website_url)
    
    # 2. If no emails/socials/whatsapps, try Exa Search & Fetch
    if not emails or not socials or not whatsapps:
        log(f"Direct scrape yielded limited results. Trying Exa for: {name}")
        search_results = exa_search(f"site:{domain} contact email phone whatsapp social", num_results=3)
        urls_to_fetch = [r.get("url") for r in search_results if r.get("url")]
        
        if urls_to_fetch:
            fetched_contents = exa_fetch(urls_to_fetch)
            for page in fetched_contents:
                text = page.get("text", "")
                
                # Extract emails
                found_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
                for email in found_emails:
                    email_lower = email.lower()
                    if not any(email_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.css', '.js']):
                        if not any(x in email_lower for x in ['wix', 'wordpress', 'example', 'domain', 'email']):
                            emails.append(email_lower)
                        
                # Extract whatsapp links/numbers
                wa_matches = re.findall(r'(?:wa\.me|api\.whatsapp\.com/send\?phone=)(\d+)', text)
                for wa in wa_matches:
                    whatsapps.append(wa)
                
                phone_matches = re.findall(r'(?:\+?974|00974|0)?[3567]\d{7}', text)
                for ph in phone_matches:
                    clean_ph = ph
                    if ph.startswith(('3', '5', '6', '7')) and len(ph) == 8:
                        clean_ph = '974' + ph
                    elif ph.startswith('0') and len(ph) == 9:
                        clean_ph = '974' + ph[1:]
                    whatsapps.append(clean_ph)
                    
                # Extract social links from text
                for platform in ['instagram', 'facebook', 'twitter', 'x.com', 'linkedin', 'youtube', 'tiktok', 'snapchat']:
                    matches = re.findall(fr'https?://(?:www\.)?{platform}\.com/([a-zA-Z0-9_.\-]+)', text)
                    for m in matches:
                        if m not in ['sharer', 'share', 'intent', 'home', 'pages']:
                            socials[platform] = f"https://{platform}.com/{m}"
                            
            emails = list(set(emails))
            whatsapps = list(set(whatsapps))
            
    # 3. Categorize emails
    general_emails = []
    marketing_emails = []
    
    marketing_keywords = ['marketing', 'media', 'advertising', 'ads', 'social', 'pr', 'commercial', 'sales', 'dعاية', 'تسويق']
    general_keywords = ['info', 'contact', 'support', 'hello', 'office', 'admin', 'connect', 'welcome']
    
    for email in emails:
        user_part = email.split('@')[0].lower()
        if any(kw in user_part for kw in marketing_keywords):
            marketing_emails.append(email)
        elif any(kw in user_part for kw in general_keywords):
            general_emails.append(email)
        else:
            general_emails.append(email)
            
    general_emails = list(set(general_emails))
    marketing_emails = list(set(marketing_emails))
    
    # 4. Get Marketing Manager LinkedIn Profiles
    marketing_linkedins = get_marketing_linkedin_profiles(name, domain)
    
    # 5. Fallback check for social links from Apollo data
    if 'linkedin_url' in company and company['linkedin_url'] and 'linkedin' not in socials:
        socials['linkedin'] = company['linkedin_url']
    if 'facebook_url' in company and company['facebook_url'] and 'facebook' not in socials:
        socials['facebook'] = company['facebook_url']
    if 'twitter_url' in company and company['twitter_url'] and 'twitter' not in socials:
        socials['twitter'] = company['twitter_url']
        
    return {
        "general_emails": general_emails,
        "marketing_emails": marketing_emails,
        "whatsapps": whatsapps,
        "socials": socials,
        "marketing_linkedins": marketing_linkedins
    }

def fetch_qatar_companies():
    log("Searching Apollo API for Qatari companies...")
    url = "https://api.apollo.io/v1/organizations/search"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY
    }
    
    # Let's search mid-sized ranges: 51-200 and 201-500
    payload = {
        "q_organization_keyword_tags": [],
        "organization_locations": ["Qatar"],
        "organization_num_employees_ranges": ["51,200", "201,500"],
        "page": 1,
        "per_page": 100
    }
    
    candidates = []
    excluded_keywords = ["real estate", "property", "properties", "developer", "estate", "عقار", "عقارات", "عقارية", "تطوير عقاري"]
    
    for page in range(1, 4): # Fetch up to 3 pages (300 companies)
        log(f"Fetching page {page} of companies...")
        payload['page'] = page
        try:
            response = requests.post(url, json=payload, headers=headers, timeout=15)
            if response.status_code != 200:
                log(f"Apollo API error: {response.status_code} - {response.text}")
                break
                
            data = response.json()
            orgs = data.get("organizations", [])
            log(f"Received {len(orgs)} organizations from page {page}")
            if not orgs:
                break
                
            for org in orgs:
                name = org.get("name", "")
                website = org.get("website_url", "")
                domain = org.get("primary_domain", "")
                industry = org.get("industry", "")
                industries = org.get("industries", [])
                
                if not website or not domain:
                    continue
                    
                # Industry filter (no real estate)
                industry_text = f"{name} {industry} {' '.join(industries)}".lower()
                if any(kw in industry_text for kw in excluded_keywords):
                    log(f"Excluding real estate company: {name}")
                    continue
                    
                candidates.append({
                    "name": name,
                    "website_url": website,
                    "primary_domain": domain,
                    "industry": industry,
                    "employees": org.get("estimated_num_employees", "غير معروف"),
                    "city": org.get("city", "غير معروف"),
                    "phone": org.get("phone", ""),
                    "linkedin_url": org.get("linkedin_url", ""),
                    "twitter_url": org.get("twitter_url", ""),
                    "facebook_url": org.get("facebook_url", "")
                })
        except Exception as e:
            log(f"Apollo connection exception on page {page}: {e}")
            break
            
    log(f"Total candidate companies collected after filtering: {len(candidates)}")
    return candidates

def main():
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
        log(f"Created export directory: {EXPORT_DIR}")
        
    candidates = fetch_qatar_companies()
    if not candidates:
        log("No candidate companies fetched. Exiting.")
        sys.exit(1)
        
    collected_leads = []
    
    # Process companies. We need exactly 30 companies with contact information.
    for idx, comp in enumerate(candidates):
        if len(collected_leads) >= 30:
            log("Reached target of 30 Qatari mid-sized companies!")
            break
            
        log(f"\nProcessing company {idx+1}/{len(candidates)}: {comp['name']}")
        
        try:
            enriched = enrich_company_details(comp)
            
            # Since the user wants real contact info and no guessing, we will require
            # that we find at least some contact detail (email, whatsapp, or socials)
            has_email = len(enriched['general_emails']) > 0 or len(enriched['marketing_emails']) > 0
            has_whatsapp = len(enriched['whatsapps']) > 0
            has_socials = len(enriched['socials']) > 0
            has_marketing_li = len(enriched['marketing_linkedins']) > 0
            
            if has_email or has_whatsapp or has_socials or has_marketing_li:
                whatsapp_str = ", ".join(enriched['whatsapps']) if enriched['whatsapps'] else comp['phone']
                
                # Prepare lead object
                lead = {
                    "اسم الشركة": comp['name'],
                    "الموقع الإلكتروني": comp['website_url'],
                    "البريد الإلكتروني العام": ", ".join(enriched['general_emails']) if enriched['general_emails'] else "غير متوفر",
                    "بريد الدعاية والتسويق": ", ".join(enriched['marketing_emails']) if enriched['marketing_emails'] else "غير متوفر",
                    "رقم الواتساب": whatsapp_str if whatsapp_str else "غير متوفر",
                    "رابط لينكد إن للشركة": enriched['socials'].get('linkedin', comp['linkedin_url'] or "غير متوفر"),
                    "حسابات مسؤولي التسويق (LinkedIn)": "\n".join(enriched['marketing_linkedins']) if enriched['marketing_linkedins'] else "غير متوفر",
                    "حساب إنستقرام": enriched['socials'].get('instagram', "غير متوفر"),
                    "حساب فيسبوك": enriched['socials'].get('facebook', "غير متوفر"),
                    "حساب تويتر/إكس": enriched['socials'].get('twitter', enriched['socials'].get('x.com', "غير متوفر")),
                    "حساب سناب شات": enriched['socials'].get('snapchat', "غير متوفر"),
                    "حساب تيك توك": enriched['socials'].get('tiktok', "غير متوفر"),
                    "حساب يوتيوب": enriched['socials'].get('youtube', "غير متوفر"),
                    "عدد الموظفين": comp['employees'],
                    "المقر (المدينة)": comp['city'],
                    "القطاع": comp['industry']
                }
                
                collected_leads.append(lead)
                log(f"Successfully added lead. Total leads collected so far: {len(collected_leads)}/30")
            else:
                log(f"Skipping {comp['name']} - insufficient contact details found.")
                
        except Exception as e:
            log(f"Error processing {comp['name']}: {e}")
            
        # Add a sleep to prevent aggressive API hitting
        time.sleep(1)
        
    if len(collected_leads) < 30:
        log(f"Warning: Only collected {len(collected_leads)} leads.")
        
    # Write to Excel
    log(f"Writing {len(collected_leads)} leads to Excel file: {OUTPUT_FILE}")
    df = pd.DataFrame(collected_leads)
    
    # Format the Excel sheet nicely
    wb = Workbook()
    ws = wb.active
    ws.title = "الشركات المتوسطة القطرية"
    
    # Set RTL layout for Arabic
    ws.views.sheetView[0].showGridLines = True
    ws.sheet_view.rightToLeft = True
    
    # Colors
    header_fill = PatternFill(start_color="005F73", end_color="005F73", fill_type="solid") # Deep Teal
    zebra_fill = PatternFill(start_color="F4F9FA", end_color="F4F9FA", fill_type="solid") # Ice Blue/Light Teal
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    title_fill = PatternFill(start_color="0A3641", end_color="0A3641", fill_type="solid") # Dark Teal
    
    # Fonts
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="333333")
    link_font = Font(name="Segoe UI", size=10, color="0000FF", underline="single")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # Write Title block
    ws.merge_cells("A1:P1")
    ws["A1"] = "إيكونوفا ستوديو - دليل بيانات الشركات القطرية المتوسطة (مستثنى العقارات)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center_align
    ws.row_dimensions[1].height = 40
    
    # Write Headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border_all
    ws.row_dimensions[2].height = 28
    
    # Write Data
    for row_idx, row_data in enumerate(df.values, 3):
        is_even = (row_idx % 2 == 0)
        current_fill = zebra_fill if is_even else white_fill
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.fill = current_fill
            cell.border = border_all
            
            # Alignments & Link Formatting
            header_name = headers[col_idx-1]
            if header_name in ["الموقع الإلكتروني", "رابط لينكد إن للشركة", "حساب إنستقرام", "حساب فيسبوك", "حساب تويتر/إكس", "حساب سناب شات", "حساب تيك توك", "حساب يوتيوب"]:
                cell.alignment = center_align
                if value != "غير متوفر" and isinstance(value, str) and value.startswith("http"):
                    cell.hyperlink = value
                    cell.font = link_font
            elif header_name in ["البريد الإلكتروني العام", "بريد الدعاية والتسويق", "رقم الواتساب", "عدد الموظفين", "المقر (المدينة)"]:
                cell.alignment = center_align
            else:
                cell.alignment = right_align
                
        ws.row_dimensions[row_idx].height = 24
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Don't size based on title row (row 1)
        for cell in col[1:]:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
                    
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    wb.save(OUTPUT_FILE)
    log("Saved Excel file successfully!")

if __name__ == "__main__":
    main()
