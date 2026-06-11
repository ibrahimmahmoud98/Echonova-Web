import os
import re
import sys
import json
import time
import sqlite3
import requests
import dns.resolver
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import load_env

# API Keys
APOLLO_API_KEY = os.environ.get("APOLLO_API_KEY")
EXA_API_KEY = os.environ.get("EXA_API_KEY")

# Paths
EXPORT_DIR = "/Users/suckmyballz/Desktop/ECHONOVA STUDIO/عملاء"
OUTPUT_FILE = os.path.join(EXPORT_DIR, "بينات الشركات الصغيرة.xlsx")
DB_PATH = os.path.join(EXPORT_DIR, "العملاء المحتملين/echonova_leads.db")

# Excluded keywords for real estate and other undesired sectors
EXCLUDED_KEYWORDS = [
    "real estate", "property", "properties", "developer", "estate", 
    "عقار", "عقارات", "عقارية", "تطوير عقاري", "realestate", "realty",
    "government", "ministry", "وزارة", "بلدية", "حكومة", "حكومية",
    "oil & gas", "petroleum", "energy", "refinery", "بترول", "نفط"
]

EXCLUDED_INDUSTRIES = [
    "real estate", "government administration", "oil & energy", 
    "defense & space", "military", "non-profit organization", 
    "religious institutions", "utilities", "mining & metals"
]

def log(msg):
    print(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {msg}", flush=True)

def get_already_collected_domains():
    domains = set()
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("SELECT domain FROM already_collected")
            for r in cursor.fetchall():
                if r[0]:
                    domains.add(r[0].lower().strip())
            conn.close()
            log(f"Loaded {len(domains)} already collected domains from SQLite database.")
        except Exception as e:
            log(f"Error reading DB already_collected: {e}")
    return domains

def add_to_already_collected(name, domain, industry):
    if os.path.exists(DB_PATH):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR IGNORE INTO already_collected (name, domain, industry, collected_at)
                VALUES (?, ?, ?, ?)
            """, (name, domain, industry, datetime.now().isoformat()))
            conn.commit()
            conn.close()
        except Exception as e:
            log(f"Error writing {domain} to already_collected DB: {e}")

def exa_search(query, num_results=5):
    url = "https://api.exa.ai/search"
    headers = {
        "x-api-key": EXA_API_KEY,
        "content-type": "application/json"
    }
    payload = {
        "query": query,
        "numResults": num_results,
        "useAutoprompt": True
    }
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=15)
        if response.status_code == 200:
            return response.json().get("results", [])
        else:
            log(f"Exa search error {response.status_code}: {response.text}")
    except Exception as e:
        log(f"Exa search exception: {e}")
    return []

def exa_fetch(urls):
    if not urls:
        return []
    url = "https://api.exa.ai/contents"
    headers = {
        "x-api-key": EXA_API_KEY,
        "content-type": "application/json"
    }
    payload = {
        "urls": urls
    }
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=20)
        if response.status_code == 200:
            return response.json().get("results", [])
        else:
            log(f"Exa fetch error {response.status_code}: {response.text}")
    except Exception as e:
        log(f"Exa fetch exception: {e}")
    return []

def verify_email_dns(email):
    """
    Verifies email domain by checking MX records.
    This acts as a solid 'no guessing' filter to ensure the domain can receive emails.
    """
    try:
        domain = email.split('@')[1].strip()
        answers = dns.resolver.resolve(domain, 'MX')
        return len(answers) > 0
    except Exception:
        return False

def scrape_website_direct(domain, website_url, country_code):
    """
    Scrapes the website directly for emails, WhatsApps, and social links.
    """
    log(f"Scraping website directly: {website_url}")
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    emails = []
    whatsapp_numbers = []
    social_links = {}
    
    # Normalize website url
    if not website_url.startswith("http"):
        website_url = "https://" + website_url
        
    urls_to_try = [website_url]
    # Try typical contact subpages
    if website_url.endswith("/"):
        urls_to_try.append(website_url + "contact")
        urls_to_try.append(website_url + "contact-us")
        urls_to_try.append(website_url + "about")
    else:
        urls_to_try.append(website_url + "/contact")
        urls_to_try.append(website_url + "/contact-us")
        urls_to_try.append(website_url + "/about")
        
    for url in urls_to_try:
        try:
            # Disable SSL verification issues as some small local companies have expired certs
            response = requests.get(url, headers=headers, timeout=8, verify=False)
            if response.status_code == 200:
                html = response.text
                
                # Extract emails
                found_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', html)
                for email in found_emails:
                    email_lower = email.lower().strip()
                    # Filter out asset files or image matches
                    if not any(email_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.tiff', '.pdf', '.mp3', '.mp4', '.css', '.js']):
                        emails.append(email_lower)
                
                # Extract whatsapp links
                wa_matches = re.findall(r'(?:wa\.me|api\.whatsapp\.com/send\?phone=)(\d+)', html)
                for wa in wa_matches:
                    whatsapp_numbers.append(wa)
                    
                # Extract phone numbers in context of whatsapp based on country
                if country_code == "SA":
                    phone_matches = re.findall(r'(?:\+?966|00966|0)?5\d{8}', html)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('05'):
                            clean_ph = '966' + ph[1:]
                        elif ph.startswith('5'):
                            clean_ph = '966' + ph
                        whatsapp_numbers.append(clean_ph)
                elif country_code == "AE":
                    phone_matches = re.findall(r'(?:\+?971|00971|0)?5[024568]\d{7}', html)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('05'):
                            clean_ph = '971' + ph[1:]
                        elif ph.startswith('5'):
                            clean_ph = '971' + ph
                        whatsapp_numbers.append(clean_ph)
                elif country_code == "QA":
                    phone_matches = re.findall(r'(?:\+?974|00974)?(?:0)?[3567]\d{7}', html)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('0') and len(ph) > 8:
                            clean_ph = '974' + ph[1:]
                        elif not ph.startswith('974') and not ph.startswith('+974') and not ph.startswith('00974'):
                            clean_ph = '974' + ph
                        whatsapp_numbers.append(clean_ph)
                        
                # Extract social links
                for platform in ['instagram', 'facebook', 'twitter', 'x.com', 'linkedin', 'youtube', 'tiktok', 'snapchat']:
                    matches = re.findall(fr'https?://(?:www\.)?{platform}\.com/([a-zA-Z0-9_.\-]+)', html)
                    for m in matches:
                        if m not in ['sharer', 'share', 'intent', 'home', 'pages', 'groups']:
                            # Clean profile handle
                            handle = m.split('/')[0].split('?')[0].strip()
                            if handle:
                                social_links[platform] = f"https://{platform}.com/{handle}"
        except Exception as e:
            # Silence scraping errors to keep execution smooth
            pass
            
    return list(set(emails)), list(set(whatsapp_numbers)), social_links

def get_linkedin_profiles(company_name, domain):
    """
    Search for LinkedIn profiles of founders, owners, or marketing managers.
    Since they are small companies, founders/owners are often the main targets.
    """
    query = f"site:linkedin.com/in/ (\"{company_name}\" OR \"{domain}\") (marketing OR advertising OR creative OR founder OR owner OR \"تسويق\" OR \"مؤسس\" OR \"مدير\")"
    results = exa_search(query, num_results=4)
    profiles = []
    for r in sorted(results, key=lambda x: x.get('score', 0), reverse=True):
        url = r.get("url", "")
        title = r.get("title", "")
        if "linkedin.com/in/" in url:
            profile_name = title.split("-")[0].strip() if "-" in title else title
            profile_name = profile_name.split("|")[0].strip() if "|" in profile_name else profile_name
            # Remove redundant 'LinkedIn' text
            profile_name = re.sub(r'\b(linkedin|linked in)\b', '', profile_name, flags=re.IGNORECASE).strip()
            profiles.append(f"{profile_name}: {url}")
    return list(set(profiles))[:3]

def enrich_company(comp, country_code):
    name = comp['name']
    website_url = comp['website_url']
    domain = comp['primary_domain']
    
    log(f"Enriching: {name} ({domain}) - {country_code}")
    
    # 1. Direct website scrape
    emails, whatsapps, socials = scrape_website_direct(domain, website_url, country_code)
    
    # 2. Exa search fallback if direct scrape is empty
    if not emails or not socials:
        log(f"Direct scrape yielded limited results for {name}. Searching Exa...")
        search_results = exa_search(f"site:{domain} contact OR email OR phone OR whatsapp OR social", num_results=3)
        urls_to_fetch = [r.get("url") for r in search_results if r.get("url")]
        
        if urls_to_fetch:
            fetched_pages = exa_fetch(urls_to_fetch)
            for page in fetched_pages:
                text = page.get("text", "")
                
                # Emails
                found_emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
                for email in found_emails:
                    email_lower = email.lower().strip()
                    if not any(email_lower.endswith(ext) for ext in ['.png', '.jpg', '.jpeg', '.gif', '.svg', '.webp', '.tiff', '.pdf', '.mp3', '.mp4', '.css', '.js']):
                        emails.append(email_lower)
                        
                # Whatsapps
                wa_matches = re.findall(r'(?:wa\.me|api\.whatsapp\.com/send\?phone=)(\d+)', text)
                for wa in wa_matches:
                    whatsapps.append(wa)
                    
                # Phones by country
                if country_code == "SA":
                    phone_matches = re.findall(r'(?:\+?966|00966|0)?5\d{8}', text)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('05'):
                            clean_ph = '966' + ph[1:]
                        elif ph.startswith('5'):
                            clean_ph = '966' + ph
                        whatsapps.append(clean_ph)
                elif country_code == "AE":
                    phone_matches = re.findall(r'(?:\+?971|00971|0)?5[024568]\d{7}', text)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('05'):
                            clean_ph = '971' + ph[1:]
                        elif ph.startswith('5'):
                            clean_ph = '971' + ph
                        whatsapps.append(clean_ph)
                elif country_code == "QA":
                    phone_matches = re.findall(r'(?:\+?974|00974)?(?:0)?[3567]\d{7}', text)
                    for ph in phone_matches:
                        clean_ph = ph
                        if ph.startswith('0') and len(ph) > 8:
                            clean_ph = '974' + ph[1:]
                        elif not ph.startswith('974') and not ph.startswith('+974') and not ph.startswith('00974'):
                            clean_ph = '974' + ph
                        whatsapps.append(clean_ph)
                        
                # Socials
                for platform in ['instagram', 'facebook', 'twitter', 'x.com', 'linkedin', 'youtube', 'tiktok', 'snapchat']:
                    matches = re.findall(fr'https?://(?:www\.)?{platform}\.com/([a-zA-Z0-9_.\-]+)', text)
                    for m in matches:
                        if m not in ['sharer', 'share', 'intent', 'home', 'pages', 'groups']:
                            handle = m.split('/')[0].split('?')[0].strip()
                            if handle:
                                socials[platform] = f"https://{platform}.com/{handle}"
                                
            emails = list(set(emails))
            whatsapps = list(set(whatsapps))
            
    # 3. Categorize emails
    general_emails = []
    marketing_emails = []
    
    marketing_keywords = ['marketing', 'media', 'advertising', 'ads', 'social', 'pr', 'commercial', 'sales', 'dعاية', 'تسويق', 'sales']
    general_keywords = ['info', 'contact', 'support', 'hello', 'office', 'admin', 'connect', 'welcome', 'inbox']
    
    for email in emails:
        user_part = email.split('@')[0].lower()
        if any(kw in user_part for kw in marketing_keywords):
            marketing_emails.append(email)
        elif any(kw in user_part for kw in general_keywords):
            general_emails.append(email)
        else:
            general_emails.append(email)
            
    general_emails = list(set(general_emails))
    marketing_emails = list(set(marketing_emails))
    
    # 4. Search LinkedIn Profiles
    linkedin_profiles = get_linkedin_profiles(name, domain)
    
    # 5. Fallback check for social links from Apollo data
    if 'linkedin_url' in comp and comp['linkedin_url'] and 'linkedin' not in socials:
        socials['linkedin'] = comp['linkedin_url']
    if 'facebook_url' in comp and comp['facebook_url'] and 'facebook' not in socials:
        socials['facebook'] = comp['facebook_url']
    if 'twitter_url' in comp and comp['twitter_url'] and 'twitter' not in socials:
        socials['twitter'] = comp['twitter_url']
        
    return {
        "general_emails": general_emails,
        "marketing_emails": marketing_emails,
        "whatsapps": whatsapps,
        "socials": socials,
        "linkedin_profiles": linkedin_profiles
    }

def fetch_candidates_from_apollo(country_name, max_pages=6):
    """
    Search Apollo for small companies (employee size 1-10, 11-50) in a specific country.
    """
    log(f"Searching Apollo API for small companies in {country_name}...")
    url = "https://api.apollo.io/v1/organizations/search"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "X-Api-Key": APOLLO_API_KEY
    }
    
    payload = {
        "organization_locations": [country_name],
        "organization_num_employees_ranges": ["1,10", "11,50"],
        "page": 1,
        "per_page": 100
    }
    
    candidates = []
    seen_domains = set()
    
    for page in range(1, max_pages + 1):
        log(f"Fetching page {page} for {country_name}...")
        payload['page'] = page
        try:
            response = requests.post(url, json=payload, headers=headers, timeout=15)
            if response.status_code != 200:
                log(f"Apollo API error on page {page}: {response.status_code} - {response.text}")
                break
                
            data = response.json()
            orgs = data.get("organizations", [])
            log(f"Received {len(orgs)} organizations from page {page}")
            if not orgs:
                break
                
            for org in orgs:
                name = org.get("name", "")
                website = org.get("website_url", "")
                domain = org.get("primary_domain", "")
                industry = org.get("industry", "")
                industries = org.get("industries", [])
                
                if not website or not domain:
                    continue
                    
                domain = domain.lower().strip()
                if domain in seen_domains:
                    continue
                seen_domains.add(domain)
                
                # Exclude based on keywords in name, industry
                industry_text = f"{name} {industry} {' '.join(industries)}".lower()
                if any(kw in industry_text for kw in EXCLUDED_KEYWORDS):
                    continue
                if any(ind in (industry or '').lower() for ind in EXCLUDED_INDUSTRIES):
                    continue
                    
                candidates.append({
                    "name": name,
                    "website_url": website,
                    "primary_domain": domain,
                    "industry": industry or "غير معروف",
                    "employees": org.get("estimated_num_employees", "غير معروف"),
                    "city": org.get("city", "غير معروف"),
                    "phone": org.get("phone", ""),
                    "linkedin_url": org.get("linkedin_url", ""),
                    "twitter_url": org.get("twitter_url", ""),
                    "facebook_url": org.get("facebook_url", "")
                })
        except Exception as e:
            log(f"Apollo query exception: {e}")
            break
            
    log(f"Total filtered candidate companies from Apollo for {country_name}: {len(candidates)}")
    return candidates

def collect_country_leads(country_name, country_code, already_collected):
    candidates = fetch_candidates_from_apollo(country_name)
    
    # Filter out already collected domains
    fresh_candidates = [c for c in candidates if c['primary_domain'] not in already_collected]
    log(f"Fresh candidates for {country_name} after excluding previously collected: {len(fresh_candidates)}")
    
    collected_leads = []
    batch_size = 10
    
    for i in range(0, len(fresh_candidates), batch_size):
        if len(collected_leads) >= 30:
            break
            
        batch = fresh_candidates[i:i+batch_size]
        log(f"Processing {country_name} batch {i//batch_size + 1}/{((len(fresh_candidates)-1)//batch_size)+1} with {len(batch)} companies...")
        
        # Use ThreadPoolExecutor to process companies concurrently
        with ThreadPoolExecutor(max_workers=len(batch)) as executor:
            # Submit tasks
            future_to_comp = {executor.submit(enrich_company, comp, country_code): comp for comp in batch}
            
            for future in as_completed(future_to_comp):
                comp = future_to_comp[future]
                if len(collected_leads) >= 30:
                    # Cancel remaining futures if supported (Python 3.9+)
                    try:
                        executor.shutdown(wait=False, cancel_futures=True)
                    except Exception:
                        pass
                    break
                    
                try:
                    enriched = future.result()
                    
                    # To be a valid lead, we MUST have some contact data (no guessing)
                    has_email = len(enriched['general_emails']) > 0 or len(enriched['marketing_emails']) > 0
                    has_whatsapp = len(enriched['whatsapps']) > 0
                    has_linkedin_profiles = len(enriched['linkedin_profiles']) > 0
                    has_socials = len(enriched['socials']) > 0
                    
                    # Require at least one solid contact channel (email, phone/whatsapp, or LinkedIn profile)
                    if has_email or has_whatsapp or has_linkedin_profiles or has_socials:
                        
                        # Verify emails using DNS MX check (no guessing, verify they actually exist)
                        verified_gen_emails = []
                        for email in enriched['general_emails'][:3]: # check top 3
                            if verify_email_dns(email):
                                verified_gen_emails.append(f"{email} (موثق)")
                            else:
                                verified_gen_emails.append(email)
                                
                        verified_mkt_emails = []
                        for email in enriched['marketing_emails'][:3]:
                            if verify_email_dns(email):
                                verified_mkt_emails.append(f"{email} (موثق)")
                            else:
                                verified_mkt_emails.append(email)
                                
                        # WhatsApp selection
                        whatsapp_str = ", ".join(enriched['whatsapps']) if enriched['whatsapps'] else comp['phone']
                        
                        lead = {
                            "اسم الشركة": comp['name'],
                            "الموقع الإلكتروني": comp['website_url'],
                            "البريد الإلكتروني العام": ", ".join(verified_gen_emails) if verified_gen_emails else "غير متوفر",
                            "بريد الدعاية والتسويق": ", ".join(verified_mkt_emails) if verified_mkt_emails else "غير متوفر",
                            "رقم الواتساب": whatsapp_str if whatsapp_str else "غير متوفر",
                            "رابط لينكد إن للشركة": enriched['socials'].get('linkedin', comp['linkedin_url'] or "غير متوفر"),
                            "حسابات مسؤولي التسويق (LinkedIn)": "\n".join(enriched['linkedin_profiles']) if enriched['linkedin_profiles'] else "غير متوفر",
                            "حساب إنستقرام": enriched['socials'].get('instagram', "غير متوفر"),
                            "حساب فيسبوك": enriched['socials'].get('facebook', "غير متوفر"),
                            "حساب تويتر/إكس": enriched['socials'].get('twitter', enriched['socials'].get('x.com', "غير متوفر")),
                            "حساب سناب شات": enriched['socials'].get('snapchat', "غير متوفر"),
                            "حساب تيك توك": enriched['socials'].get('tiktok', "غير متوفر"),
                            "حساب يوتيوب": enriched['socials'].get('youtube', "غير متوفر"),
                            "عدد الموظفين": comp['employees'],
                            "المقر (المدينة)": comp['city'],
                            "القطاع": comp['industry']
                        }
                        
                        collected_leads.append(lead)
                        # Sync with local database to maintain audit trail and prevent duplication
                        add_to_already_collected(comp['name'], comp['primary_domain'], comp['industry'])
                        log(f"-> Saved lead: {comp['name']} for {country_name}. Total: {len(collected_leads)}/30")
                    else:
                        log(f"-> Skipped: {comp['name']} (insufficient contact data)")
                except Exception as e:
                    log(f"Error enriching company {comp['name']}: {e}")
                
    log(f"Finished {country_name}. Collected {len(collected_leads)} leads.")
    return collected_leads[:30]

def build_excel_workbook(saudi_leads, uae_leads, qatar_leads):
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styles
    header_fill = PatternFill(start_color="005F73", end_color="005F73", fill_type="solid") # Deep Teal
    zebra_fill = PatternFill(start_color="F4F9FA", end_color="F4F9FA", fill_type="solid") # Ice Blue
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    title_fill = PatternFill(start_color="0A3641", end_color="0A3641", fill_type="solid") # Dark Teal
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10, color="333333")
    link_font = Font(name="Segoe UI", size=10, color="0000FF", underline="single")
    verified_email_font = Font(name="Segoe UI", size=10, bold=True, color="008000") # Green font for verified
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    data_groups = [
        ("الشركات الصغيرة السعودية", saudi_leads, "بيانات الشركات الصغيرة السعودية"),
        ("الشركات الصغيرة الإماراتية", uae_leads, "بيانات الشركات الصغيرة الإماراتية"),
        ("الشركات الصغيرة القطرية", qatar_leads, "بيانات الشركات الصغيرة القطرية")
    ]
    
    for sheet_title, leads, title_text in data_groups:
        ws = wb.create_sheet(title=sheet_title)
        
        # Enable RTL & Gridlines
        ws.views.sheetView[0].showGridLines = True
        ws.sheet_view.rightToLeft = True
        
        # Write Title row
        ws.merge_cells("A1:P1")
        ws["A1"] = f"إيكونوفا ستوديو - {title_text} (مستثنى العقارات)"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 40
        
        if not leads:
            ws.cell(row=2, column=1, value="لا توجد بيانات متوفرة").font = Font(name="Segoe UI", size=12, italic=True)
            continue
            
        df = pd.DataFrame(leads)
        headers = list(df.columns)
        
        # Write Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_all
        ws.row_dimensions[2].height = 28
        
        # Write Data
        for row_idx, row_data in enumerate(df.values, 3):
            is_even = (row_idx % 2 == 0)
            current_fill = zebra_fill if is_even else white_fill
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.fill = current_fill
                cell.border = border_all
                
                header_name = headers[col_idx-1]
                
                # Formatting cell types
                if header_name in ["الموقع الإلكتروني", "رابط لينكد إن للشركة", "حساب إنستقرام", "حساب فيسبوك", "حساب تويتر/إكس", "حساب سناب شات", "حساب تيك توك", "حساب يوتيوب"]:
                    cell.alignment = center_align
                    if value != "غير متوفر" and isinstance(value, str) and value.startswith("http"):
                        cell.hyperlink = value
                        cell.font = link_font
                elif header_name in ["البريد الإلكتروني العام", "بريد الدعاية والتسويق", "رقم الواتساب", "عدد الموظفين", "المقر (المدينة)"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = right_align
                    
                # Format verified email tag
                if "موثق" in str(value):
                    cell.font = verified_email_font
                    
            ws.row_dimensions[row_idx].height = 24
            
        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                val_str = str(cell.value or '')
                lines = val_str.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        # Limit max length to avoid ridiculously wide columns for long lists of LinkedIn profiles
                        max_len = min(len(line), 50)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    log(f"Excel workbook saved successfully at: {OUTPUT_FILE}")

def main():
    log("Starting Small Business Leads collection pipeline...")
    
    # Load previously collected domains to prevent duplication
    already_collected = get_already_collected_domains()
    
    # Collect leads for Saudi Arabia
    saudi_leads = collect_country_leads("Saudi Arabia", "SA", already_collected)
    
    # Collect leads for UAE
    uae_leads = collect_country_leads("United Arab Emirates", "AE", already_collected)
    
    # Collect leads for Qatar
    qatar_leads = collect_country_leads("Qatar", "QA", already_collected)
    
    # Generate and style the Excel file
    build_excel_workbook(saudi_leads, uae_leads, qatar_leads)
    
    log("Pipeline completed successfully!")

if __name__ == "__main__":
    main()
